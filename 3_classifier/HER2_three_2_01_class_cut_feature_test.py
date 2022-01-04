import h5py
import sklearn.metrics as metrics
from openpyxl import workbook
from sklearn.datasets import load_iris
from sklearn.model_selection import train_test_split, KFold, StratifiedKFold
from sklearn.svm import SVC
from sklearn.model_selection import GridSearchCV
import os
import gc
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, AdaBoostClassifier
from sklearn.decomposition import PCA, KernelPCA
from sklearn.naive_bayes import MultinomialNB
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
import copy
from scipy import stats
from statsmodels.stats.diagnostic import lilliefors
from sklearn import preprocessing
import umap
import warnings
from sklearn.cluster import KMeans, SpectralClustering
import pandas as pd
import random
import numpy as np
import pickle

warnings.filterwarnings('ignore')
from sklearn.linear_model import LassoCV
from sklearn.linear_model import Lasso
from sklearn import feature_selection



sep = os.sep
filesep = sep

# 剔除肿块
zhongkai_task = False
zhongkai_task_out = False  # 只剔除外部的非肿块
Nzhongkai_task_out = False

task = 'label_her2'  # label_bc  label_tils  label_her2(零表达 低表达)  'label_her2_fenxing'（阴0 阳1）

# 设置随机数种子
random.seed(6666)
# dsfr radiomics dsfr+radiomics

'''
1:dsfr
2:radiomics
3:'dsfr+radiomics'
4:dsfr  radiomics分数相加
5.dsfr  dsfr+radiomics分数相加
6.radiomics  dsfr+radiomics分数相加
7 dsfr radiomics  dsfr+radiomics 分数相加
'''

key_word_model = ['radiomics']

Inner_PATIENTS_Label = {}
Out_PATIENTS_Label = {}
Root_Path = 'E:\workspace\BC\github'
####################得到label###
pkl_file = open(os.path.join(Root_Path,'all_label.pkl'), 'rb')
alldata_label = pickle.load(pkl_file)
in_p_num = 0
out_p_num = 0
sfy2_no_id = [35151,37555,41868,43718,40131]
sfy_no_id = [73513,64777]
sy_no_id = [203938,397523]
new_sy = [78651,60907,70684]
for pth_id in alldata_label.keys():
    if pth_id in sy_no_id+sfy2_no_id +sfy_no_id:
        continue
    if zhongkai_task:
        if alldata_label[pth_id]['label_zk'] != 1:
            continue

    if alldata_label[pth_id]['set'] == 'sy':
        if alldata_label[pth_id][task] < 3:
            if alldata_label[pth_id][task]==0 or alldata_label[pth_id][task]==1:
                Inner_PATIENTS_Label[pth_id] = 1
            else:
                Inner_PATIENTS_Label[pth_id] = 0
            if Inner_PATIENTS_Label[pth_id] == 1:
                in_p_num = in_p_num + 1

    elif alldata_label[pth_id]['set'] == 'sfy':
        if alldata_label[pth_id][task] < 3:
            if zhongkai_task_out:
                if alldata_label[pth_id]['label_zk'] != 1:
                    continue
            if Nzhongkai_task_out:
                if alldata_label[pth_id]['label_zk'] == 1:
                    continue
            if alldata_label[pth_id][task] ==0 or alldata_label[pth_id][task] ==1:
                Out_PATIENTS_Label[pth_id] = 1
            else:
                Out_PATIENTS_Label[pth_id] = 0
            if Out_PATIENTS_Label[pth_id] == 1:
                out_p_num = out_p_num + 1



# 在任意轴打乱顺序
def array_shuffle(x, axis=0):
    new_index = list(range(x.shape[axis]))
    random.shuffle(new_index)
    x_new = np.transpose(x, ([axis] + [i for i in list(range(len(x.shape))) if i is not axis]))
    x_new = x_new[new_index][:]
    new_dim = list(np.array(range(axis)) + 1) + [0] + list(np.array(range(len(x.shape) - axis - 1)) + axis + 1)
    x_new = np.transpose(x_new, tuple(new_dim))
    return x_new


def sigmoid_y(x, thresold=0.5):
    if x < thresold:
        x = 0
    else:
        x = 1
    return x


def getAccSenSpcAuc(label, pre, pre_bestthresold=None):
    """
    只适用于二分类
    :param label:01标签
    :param pre:属于1类的概率
    :param pre_bestthresold:可手动设置阈值，否则返回roc曲线上约登指数最大处阈值
    :return:
    """
    final_true_label = label
    final_pred_value = pre
    patient_num = len(final_true_label)

    # 计算auc，并计算最佳阈值
    if (sum(final_true_label) == patient_num) or (sum(final_true_label) == 0):
        Aucc = 0
        print('only one class')
    else:
        Aucc = metrics.roc_auc_score(final_true_label, final_pred_value)
        # print('AUC', Aucc)

    # 计算最佳阈值
    fpr, tpr, thresholds = metrics.roc_curve(final_true_label, final_pred_value)
    # 计算约登指数
    Youden_index = tpr + (1 - fpr)
    best_thresold = thresholds[Youden_index == np.max(Youden_index)][0]

    # have no idea about that threshold is bigger than 1 sometimes
    # maybe can find in https://github.com/scikit-learn/scikit-learn/commit/4d9a67f77787ffe9955187865f9b95e19286f069
    # or https://github.com/scikit-learn/scikit-learn/issues/3097
    if best_thresold > 1:
        best_thresold = 0.5

    # best_thresold = 0.5  # 认为设定阈值，因为如果预测值很hard（偏向0或1），那么阈值也会很偏（毕竟取的是预测值），但是实际情况可能取0.5为阈值也没啥大区别

    # 如果有预设阈值，则使用预设阈值计算acc，sen，spc
    if pre_bestthresold is not None:
        best_thresold = pre_bestthresold

    # 根据最终list来计算最终指标
    tp = 0
    tn = 0
    fp = 0
    fn = 0

    for nn in range(patient_num):
        t_label = final_true_label[nn]  # true label
        p_value = final_pred_value[nn]

        p_label = sigmoid_y(p_value, best_thresold)

        if (t_label == 1) and (t_label == p_label):
            tp = tp + 1  # 真阳
        elif (t_label == 0) and (t_label == p_label):
            tn = tn + 1  # 真阴
        elif (t_label == 1) and (p_label == 0):
            fn = fn + 1  # 假阴
        elif (t_label == 0) and (p_label == 1):
            fp = fp + 1  # 假阳

    Sensitivity = tp / ((tp + fn) + (1e-16))
    Specificity = tn / ((tn + fp) + (1e-16))
    Accuracy = (tp + tn) / ((tp + tn + fp + fn) + (1e-16))

    return [Accuracy, Sensitivity, Specificity, Aucc, best_thresold]


# 数据选择
# 正态分布测试
def check_normality(testData, printflag=False):
    # 20<样本数<50用normal test算法检验正态分布性
    if 20 < len(testData) < 50:
        if printflag:
            print('use normal test')
        p_value = stats.normaltest(testData)
        return [p_value[0], p_value[1]]

    # 样本数小于50用Shapiro-Wilk算法检验正态分布性
    if len(testData) < 50:
        if printflag:
            print('use shapiro test')
        p_value = stats.shapiro(testData)
        return [p_value[0], p_value[1]]

    if 300 >= len(testData) >= 50:
        if printflag:
            print('use lilliefors test')
        p_value = lilliefors(testData)
        return [p_value[0], p_value[1]]

    if len(testData) > 300:
        if printflag:
            print('use kstest test')
        p_value = stats.kstest(testData, 'norm')
        return [p_value[0], p_value[1]]


# 　输入为训练集数据和标签，测试集数据，特征个数
def find_feature_Ttest(Data, label, thresold=0.0495):
    zeroflag = 0
    oneflage = 0
    length = Data.shape[1]
    for i in range(0, len(label)):
        if label[i] == 0:
            zeroflag = zeroflag + 1
        else:
            oneflage = oneflage + 1
    # print('nagative:', zeroflag, 'positive', oneflage)

    nor_num = 0
    ttest_num = 0
    ttest_nor = 0
    num = 0
    feature_label = np.zeros([1, length], dtype=np.int)

    for i in range(length):
        testdata1 = Data[:, i][label == 0]
        testdata2 = Data[:, i][label == 1]
        ksresult1 = check_normality(testdata1)
        ksresult2 = check_normality(testdata2)

        if ((ksresult1[1] > 0.05) and (ksresult2[1] > 0.05)):
            # if False:
            num = num + 1
            leveneresult = stats.levene(testdata1, testdata2)
            #        print(leveneresult)
            if leveneresult[1] >= 0.05:
                ttestresult = stats.ttest_ind(testdata1, testdata2)
                #            print(ttestresult)
                if ttestresult[1] < thresold:
                    # train_data.append(Data[:, i])
                    ttest_num = ttest_num + 1
                    feature_label[0][i] = 1

                    #            ttest
            if leveneresult[1] < 0.05:
                ttestresult = stats.ttest_ind(testdata1, testdata2, equal_var=False)
                if ttestresult[1] < thresold:
                    ttest_nor = ttest_nor + 1
                    feature_label[0][i] = 1
                    # train_data.append(Data[:, i])
                    #                print(ttestresult)
        # else:
        #     nontestresult = stats.mannwhitneyu(testdata1, testdata2)
        #     if nontestresult[1] < thresold:
        #         feature_label[0][i] = 1
        #         # train_data.append(Data[:, i])
        #         nor_num = nor_num + 1

    # print('check_normality:', num)
    # print('nor_num',nor_num)
    # print('ttest_num', ttest_num)
    # print('ttest_nor', ttest_nor)
    # print('nontestresult',nontestresult)
    return feature_label


# find_feature_Ttest(feature_train, label_train)


def zhenghe_xxt(data, distill_mode):
    """
    # 整理数据到病人单位 ，即将病人特征储存到字典里===============================================
    :param data: 字典
    :param distill_mode: 三个,max,mean,maxmean

    :return:
    """
    id_all = []
    all_key = data.keys()  # np.unique会自动升序排序
    data_dict = {}
    for id in all_key:
        # data_person = data[data[:, 0] == id, :]
        data_dict[id] = {'dsfrfeature': data[id]['dsfr_feature'], 'radiofeature': data[id]['radio_feature'],
                         'label': data[id]['label']}

    # 提前整合每个病人的特征(取均值or max，or均值+max titu1996那个方法) =======================================
    # in
    feature = []
    radiofeature = []
    label = []
    for idd in all_key:

        if distill_mode is 'max':
            # print('distill_mode is ', distill_mode)
            # data_dict[idd]['dsfrfeature'] = list(data_dict[idd]['dsfrfeature'])
            feature.append(list(np.max(data_dict[idd]['dsfrfeature'], axis=0)))  # 列
        elif distill_mode is 'mean':
            # print('distill_mode is ', distill_mode)
            feature.append(list(np.mean(data_dict[idd]['dsfrfeature'], axis=0)))
        elif distill_mode is 'maxmean':
            # print('distill_mode is ', distill_mode)
            feature.append(list(0.5 * (np.mean(data_dict[idd]['dsfrfeature'], axis=0) +
                                       np.max(data_dict[idd]['dsfrfeature'], axis=0))))
        else:
            raise Exception("distill_mode must be one of them (max,mean,maxmean) ")
        radiofeature.append(list(data_dict[idd]['radiofeature']))
        id_all.append(int(idd))
        label.append(data_dict[idd]['label'])
    dsfr_feature = np.array(feature)
    radio_feature = np.array(radiofeature)
    label = np.array(label)
    id_all = np.array(id_all)
    return [id_all, label, dsfr_feature, radio_feature]


def sub_undersampled_use_clustering(feature_train, label_train, train_id, add_num=5):
    # feature_train, label_train, current_fold_train_id
    """
    使用聚类方法进行样本欠采样
    (适用于多分类)
    :return:
    """
    # 先拼接起来
    label_train = np.reshape(label_train, [len(label_train), 1])
    train_id = np.array(train_id, dtype=np.float)
    train_id = np.reshape(train_id, [len(train_id), 1])
    d_data = np.concatenate((train_id, label_train, feature_train), axis=1)

    # 按照label列排序，把每一类样本分开处理
    d_data = d_data[d_data[:, 1].argsort()]

    class_name = np.unique(d_data[:, 1])  # 类别的名字
    class_num = len(class_name)  # 一共有几个类别
    class_per_sub_num = [sum(d_data[:, 1] == t_classes) for t_classes in class_name]  # 每类共包括多少样本
    hold_class = class_name[class_per_sub_num.index(min(class_per_sub_num))]  # 需要被保护，不欠采样的样本(这里的index只会返回第一个，哪怕有相同值)
    print(class_name)
    """
    注意，多分类与二分类，欠采样的情况，有多种
    若两类样本数量相同，则不进行欠采样
    """
    # 分离出被保护类别的样本
    hold_data = d_data[d_data[:, 1] == hold_class, :]
    hold_class_per_sub_num = hold_data.shape[0]
    final_decrease_num = add_num + hold_class_per_sub_num
    # ps：需要人为调整一下最终欠采样的个数，但是还是应该用一种自适应的方法比较好

    # 依次处理需要被欠采样的类别，并压入一个list（最后再把这个list合并在一起返回）
    # 如果数量与被保护的不相等，那么就进行欠采样，否则直接压入
    data_after_decrease = []
    for c_index, t_class_nume in enumerate(class_name):
        if t_class_nume != hold_class:
            # if class_per_sub_num[c_index] != hold_class_per_sub_num:
            print(t_class_nume, c_index)
            decrease_data = d_data[d_data[:, 1] == t_class_nume, :]

            # 聚类（暂时先聚类为2类）
            # n_cluster = KMeans(n_clusters=2, random_state=0).fit_predict(decrease_data[:, 2:])  # 使用Kmeans聚类
            n_cluster = SpectralClustering(n_clusters=2, gamma=0.5, random_state=0).fit_predict(
                decrease_data[:, 2:])  # 使用谱聚类
            big_cluster = np.argmax(np.bincount(n_cluster))  # 最大簇的索引号
            small_cluster = np.argmin(np.bincount(n_cluster))  # 较小簇的索引号

            big_cluster_data = decrease_data[n_cluster == big_cluster, :]
            small_cluster_data = decrease_data[n_cluster == small_cluster, :]

            big_num = np.max(np.bincount(n_cluster))
            small_num = np.max(np.bincount(n_cluster))

            # 之后就分情况，进行欠采样
            if big_num >= final_decrease_num:
                big_cluster_data = array_shuffle(big_cluster_data, axis=0)
                big_cluster_data = big_cluster_data[:final_decrease_num, :]
                data_after_decrease.append(big_cluster_data)
            else:
                res_num = final_decrease_num - big_num
                small_cluster_data = array_shuffle(small_cluster_data, axis=0)
                small_cluster_data = small_cluster_data[:res_num, :]
                big_cluster_data = np.concatenate((big_cluster_data, small_cluster_data), axis=0)
                data_after_decrease.append(big_cluster_data)
        # else:
        #     decrease_data = d_data[d_data[:, 1] == t_class_nume, :]
        #     data_after_decrease.append(decrease_data)
    if len(data_after_decrease) > 1:
        final_data = np.concatenate(data_after_decrease, axis=0)
    else:
        final_data = data_after_decrease[0]

    final_data = np.concatenate((final_data, hold_data), axis=0)

    final_feature = final_data[:, 2:]
    final_label = np.reshape(final_data[:, 1], final_data.shape[0])
    final_id = list(final_data[:, 0])

    return [final_feature, final_label, final_id]


def find_feature_lassocv(Data, label):
    alphas = np.logspace(-3, 1, 50)  # 45
    # alphas = np.logspace(-100, 1, 50)
    model_lassoCV = LassoCV(alphas=alphas, cv=10, max_iter=600).fit(Data, label)
    # model_lassoCV = Lasso(alpha=0.003,max_iter=500).fit(Data, label)  # HER2 #a=0.001 rd = 26 AUC 0.8
    # model_lassoCV = LassoCV(cv=10,max_iter=500,n_alphas=100,n_jobs=15,tol=1e-6).fit(Data, label)
    coef = model_lassoCV.coef_
    return coef


def find_feature_lasso(Data, label, alpha,random_state):
    # alphas = np.logspace(-100, 1, 50)
    # model_lassoCV = LassoCV(alphas=alphas, cv=10,max_iter=1000).fit(Data, label)
    model_lassoCV = Lasso(alpha=alpha, max_iter=500,random_state=random_state).fit(Data, label)  # HER2 #0.0008 #a=0.001 rd = 26 AUC 0.8
    # model_lassoCV = LassoCV(cv=10,max_iter=500,n_alphas=100,n_jobs=15,tol=1e-6).fit(Data, label)
    coef = model_lassoCV.coef_
    return coef


def choose_feature(feature,label,feature_out=None,xuhao=None,select_first=True, usepca=False, usekernelpca=False, uselda=False,
                   useUmap=False,
                   usestatics=False, uselasso=False, alpha=0.0008, uselassocv=False, useSelectPercentile=False,
                   SelectPercentile_fc=2, pers_chi=5,random_state=None):
    # 特征降维 ------------------------------------------------------------------------------------------
    # ===============================================参数设定 =================================================================
    test_outside = False
    if feature_out is not None:
        test_outside = True
    xuhao_out = 0
    select_first = select_first
    # lasso类
    uselasso = uselasso
    uselassocv = uselassocv
    alpha = alpha
    # pca类
    pcara = 0.98  # pca
    usepca = usepca
    usekernelpca = usekernelpca
    kernel_n_components = 25

    # lda
    uselda = uselda

    # 统计方法
    usestatics = usestatics

    # umap
    useUmap = useUmap
    umap_n_components = 40
    umap_n_neighbors = 12
    umap_min_dist = 0.001
    umap_n_epochs = 350

    print('select features first')
    # ===============================================特征筛选 =================================================================
    if useSelectPercentile:
        '''         f_classif : ANOVA F-value between label/feature for classification tasks.分类任务中标签/特征间的方差分析f值
                    mutual_info_classif : Mutual information for a discrete target.          离散目标的互信息。
                    chi2 : Chi-squared stats of non-negative features for classification tasks. 分类任务非负特征的卡方统计。
                    f_regression : F-value between label/feature for regression tasks.       用于回归任务的标签/特征之间的f值。
                    mutual_info_regression : Mutual information for a continuous target.    连续目标的互信息。
                    SelectKBest : Select features based on the k highest scores.          根据k个最高分选择特征。
                    SelectFpr : Select features based on a false positive rate test.          根据假阳性率测试选择特征。  
                    SelectFdr : Select features based on an estimated false discovery rate.    根据估计的错误发现率选择特征。
                    SelectFwe : Select features based on family-wise error rate.              根据家族错误率选择特征。
                    GenericUnivariateSelect : Univariate feature selector with configurablemode.   具有可配置的单变量特性选择器
        '''
        pers_chi = pers_chi  # 是筛选后剩下的特征百分比
        if SelectPercentile_fc == 1:
            function_choose = feature_selection.f_classif  # 筛选的方式
        elif SelectPercentile_fc == 2:
            function_choose = feature_selection.mutual_info_classif
        elif SelectPercentile_fc == 3:
            function_choose = feature_selection.chi2
        elif SelectPercentile_fc == 4:
            function_choose = feature_selection.f_regression
        elif SelectPercentile_fc == 5:
            function_choose = feature_selection.mutual_info_regression
        elif SelectPercentile_fc == 6:
            function_choose = feature_selection.SelectKBest
        elif SelectPercentile_fc == 7:
            function_choose = feature_selection.SelectFpr
        elif SelectPercentile_fc == 8:
            function_choose = feature_selection.SelectFdr
        elif SelectPercentile_fc == 9:
            function_choose = feature_selection.SelectFwe
        elif SelectPercentile_fc == 10:
            function_choose = feature_selection.GenericUnivariateSelect

        fs = feature_selection.SelectPercentile(function_choose, percentile=int(pers_chi))
        feature = fs.fit_transform(feature, label)

        if test_outside:
            feature_out = fs.transform(feature_out)
            xuhao_out = fs.transform(xuhao)

    # pca
    if usepca:
        pca = PCA(pcara)
        feature = pca.fit_transform(feature)
        if test_outside:
            feature_out = pca.transform(feature_out)

    # lda 注意，只能降到1到K-1的范围，K是类别数
    if uselda:
        lda = LinearDiscriminantAnalysis(n_components=1)
        feature = lda.fit_transform(feature, label)
        if test_outside:
            feature_out = lda.transform(feature_out)

    # 统计方法（只适用于2分类）
    if usestatics:
        feature_index = find_feature_Ttest(feature, label)
        print('significant feature num:', np.sum(feature_index))
        feature_index = np.squeeze(feature_index)
        if np.sum(feature_index) == 0:
            # 如果一个显著的都没有，那么就取全部的特征
            feature_index = (feature_index + 1)
            feature_index = feature_index / feature_index
        feature = feature[:, feature_index == 1]
        if test_outside:
            feature_out = feature_out[:, feature_index == 1]
    # lasso筛选特征
    if uselasso:
        coel = find_feature_lasso(feature, label, alpha,random_state=random_state)
        feature = feature[:, coel != 0]
        if test_outside:
            feature_out = feature_out[:, coel != 0]
            xuhao_out = xuhao[:, coel != 0]

    if uselassocv:
        coel = find_feature_lassocv(feature, label)
        feature = feature[:, coel != 0]
        if test_outside:
            feature_out = feature_out[:, coel != 0]
    # if uselasso:
    #     lasso = Lasso(1000)  # 调用Lasso()函数，设置λ的值为1000
    #     lasso.fit(feature, label)
    #     print('相关系数为：', np.round(lasso.coef_, 5))  # 输出结果，保留五位小数
    #     ## 计算相关系数非零的个数
    #     print('相关系数非零个数为：', np.sum(lasso.coef_ != 0))
    #     mask = lasso.coef_ != 0  # 返回一个相关系数是否为零的布尔数组
    #     print('相关系数是否为零：', mask)
    #     new_reg_data = data.iloc[:, mask]  # 返回相关系数非零的数据
    #     new_reg_data.to_csv(outputfile)  # 存储数据
    #     print('输出数据的维度为：', new_reg_data.shape)  # 查看输出数据的维度

    # 流形学习：Umap
    if useUmap:
        umap_dr = umap.UMAP(n_components=umap_n_components, n_neighbors=umap_n_neighbors,
                            min_dist=umap_min_dist, n_epochs=umap_n_epochs)
        feature = umap_dr.fit_transform(feature)
        if test_outside:
            feature_out = umap_dr.transform(feature_out)

    # 核pca
    if usekernelpca:
        k_pca = KernelPCA(n_components=kernel_n_components, kernel='rbf')
        feature = k_pca.fit_transform(feature, label)
        # feature_train = (feature_train - feature_train.min(axis=0)) / (feature_train.max(axis=0) - feature_train.min(axis=0) + 1e-12)
        # feature_test = (feature_test - feature_train.min(axis=0)) / (feature_train.max(axis=0) - feature_train.min(axis=0) + 1e-12)
        if test_outside:
            feature_out = k_pca.transform(feature_out)

    # 后蛟的方法（暂略）
    return feature, feature_out,xuhao_out


# 至此，特征降维结束 ======================================================================================

if __name__ == '__main__':
    use_subject_decrease = False  # 是否使用欠采样
    decrease_method = 3  # 暂时只实现了方法3，所以这个参数先不用改
    """
    欠采样分为以下几种具体情况（包括以后的过采样）
    （1）只在特征选择中应用欠采样：
         在每一折中，先对样本欠采样，之后特征选择，之后再把丢弃的样本拿回来建模
    （2）只在构建分类器中应用欠采样：
         在每一折中，先进行特征选择，之后对样本欠采样来建模
    （3）同时在特征选择和建模过程中应用欠采样：
         在每一折中，先对样本欠采样，之后特征选择并直接建模，不放回样本    

    ps：目前暂时先实现功能（3）
    """
    # ====================设定参数 ===================================================================

    # ===================保存路径和输入特征路径====================================================================

    savepath = os.path.join(Root_Path,'result', task, '2_elseresult')  # 结果保存路径
    if not os.path.exists(savepath):
        os.makedirs(savepath)

    icc_excel = pd.read_excel(os.path.join(Root_Path,'all_ICC59.xlsx'))#记录了组学特征的ICC值，并从小到大排列
    inner_radio_feature_path = os.path.join(Root_Path,'feature','inner_feature','radiomics')
    inner_dsfr_feature_path = os.path.join(Root_Path,'feature','inner_feature','dsfr')
    out_radio_feature_path = os.path.join(Root_Path,'feature','out_feature','radiomics')
    out_dsfr_feature_path = os.path.join(Root_Path,'feature','out_feature','dsfr')

    # 其他参数 -----------------------------------------StratifiedKFold-----------------------------------------------------
    test_outside = True
    holdout = False  # 是否留一交叉验证,false则为k_fold折
    k_fold = 10  # 交叉验证折数，带平衡！！的交叉验证分折
    # in : 6-fold,SSTR2 16-fold VEGFR2, 11-fold MGMT
    # out: 2-fold,SSTR2 7-fold VEGFR2, 10-fold MGMT
    N = 1  # 跑N次重复试验

    # 是否设定随机参数网格（这个不是随机寻优，只是随机赋参数跑,类似抽奖，只用在预实验，标准实验不这么搞）
    random_grid = True
    state = 'random_grid2'
    # 是否网格寻优       （标准实验这么搞！！，如果使用网格寻优，则会覆盖掉random_grid的设置，详见代码调用）
    grid_search = False
    show_grid = False  # 是否打印网格寻优细节
    gridcv = 4
    if grid_search:
        state = 'grid_search'

    distill_mode = 'mean'  # 病人block整合为一个向量的策略，max，mean或者maxmean
    save_prob = True  # 是否保留每次实验的预测值（if true，prob，label和id都要保存到一个txt中）

    # 是否使用某种分类器
    UseLR = True  # 是否使用RF

    # 外部检查点=====================================================================
    # 建立个文件，代表开始run
    filename = savepath + sep + 'start_flagfile.txt'
    f = open(filename, 'a+')
    writewords = 'its running !!!!!!'
    f.write(writewords)
    f.close()

    # ============================================================================
    # ============================================================================
    # ========         =====    =====  ===    =====  =============================
    # ========  ======  ====  =  ====  ===  =  ====  =============================
    # ========  ======  ====  ==  ===  ===  ==  ===  =============================
    # ========      ========  ===  ==  ===  ===  ==  =============================
    # ========  ===   ======  ====  =  ===  ====  =  =============================
    # ========  ======   ===  =====    ===  =====    =============================
    # ============================================================================
    # ============================================================================
    # 大循环，重复试验
    # featurepathlist = get_filelist_frompath(featurepath, 'h5')
    inner_radio_feature_path_list = os.listdir(inner_radio_feature_path)
    inner_dsfr_feature_path_list = os.listdir(inner_dsfr_feature_path)
    out_radio_feature_path_list = os.listdir(out_radio_feature_path)
    out_dsfr_feature_path_list = os.listdir(out_dsfr_feature_path)

    test1 = 0
    for iiiiiiiiii in range(N):
        gc.collect()  # 释放内存
        print(iiiiiiiiii)
        # 每轮跑完更新一下列表，取结果较好的几个再调整参数(代码后面再写）
        # for datapath in featurepathlist:
        for xunhuan in range(1):
            test1 = test1 + 1
            print('test:', test1, '/', N, '======================================================')
            # print(datapath)

            # 读数据 ==================================================================================================
            # 读数据step1：首先读入数据
            data = {}
            for pth_id in Inner_PATIENTS_Label.keys():
                # dsfr_feature
                H5_file_dsfr = h5py.File(os.path.join(inner_dsfr_feature_path, str(pth_id) + '.h5'), 'r')
                pth_name = int(pth_id)
                # data1 = H5_file1['Kmeans_max_mask_f_value'][:]
                #dsfr_feature = H5_file_dsfr['f_values'][:]
                dsfr_feature = H5_file_dsfr['Kmeans_max_mask_f_value'][:]
                # dsfr_feature = H5_file_dsfr['Features'][:]
                H5_file_dsfr.close()
                data[int(pth_name)] = {}
                data[int(pth_name)]['dsfr_feature'] = dsfr_feature
                data[int(pth_name)]['label'] = Inner_PATIENTS_Label[int(pth_name)]
                # radio_feature
                pkl_file = open(os.path.join(inner_radio_feature_path, str(pth_id) + '.pkl'), 'rb')
                Features_message = pickle.load(pkl_file)
                feature = Features_message['Features']
                if len(feature.shape) == 1:
                    radio_feature = feature
                else:
                    radio_feature = feature[0]
                pkl_file.close()
                data[int(pth_name)]['radio_feature'] = radio_feature

            # 读数据step2：读入外验证数据
            out_data = {}
            if test_outside:
                for pth_id in Out_PATIENTS_Label.keys():
                    pth_name = int(pth_id)
                    # dsfr
                    H5_file = h5py.File(os.path.join(out_dsfr_feature_path, str(pth_id) + '.h5'), 'r')
                    out_dsfr_feature = H5_file['Kmeans_max_mask_f_value'][:]
                    H5_file.close()

                    out_data[int(pth_name)] = {}
                    out_data[int(pth_name)]['dsfr_feature'] = out_dsfr_feature
                    out_data[int(pth_name)]['label'] = Out_PATIENTS_Label[int(pth_name)]
                    # radiomics

                    pkl_file = open(os.path.join(out_radio_feature_path, str(pth_id) + '.pkl'), 'rb')
                    Features_message = pickle.load(pkl_file)
                    feature = Features_message['Features']
                    if len(feature.shape) == 1:
                        out_radio_feature = feature
                    else:
                        out_radio_feature = feature[0]

                    out_data[int(pth_name)]['radio_feature'] = out_radio_feature
                    H5_file.close()

            # 整理数据到病人单位，即将病人特征储存到字典里===============================================
            id_all, label, dsfrfeature, radiofeature = zhenghe_xxt(data, distill_mode)
            # id_all, label, feature = [int(pth_name),Inner_PATIENTS_LABEL[int(pth_name)],data]
            print(
                '{} in data num    2:1+0=={}:{}'.format(len(id_all), int(len(Inner_PATIENTS_Label) - in_p_num), in_p_num))

            if test_outside:
                id_all_out, label_out, dsfrfeature_out, radiofeature_out = zhenghe_xxt(out_data, distill_mode)
                print('{} out data num    2:1+0=={}:{}'.format(len(id_all_out), int(len(Out_PATIENTS_Label) - out_p_num),
                                                             out_p_num))

            # 特征归一化
            # 先归一化外部，不然内部被覆盖后，最值就变了
            if test_outside:  # 按照内部数据归一化
                radiofeature_out = (radiofeature_out - radiofeature.min(axis=0)) / (
                            radiofeature.max(axis=0) - radiofeature.min(axis=0) + 1e-12)
                dsfrfeature_out = (dsfrfeature_out - dsfrfeature.min(axis=0)) / (
                            dsfrfeature.max(axis=0) - dsfrfeature.min(axis=0) + 1e-12)

            dsfrfeature = (dsfrfeature - dsfrfeature.min(axis=0)) / (
                        dsfrfeature.max(axis=0) - dsfrfeature.min(axis=0) + 1e-12)
            radiofeature = (radiofeature - radiofeature.min(axis=0)) / (
                        radiofeature.max(axis=0) - radiofeature.min(axis=0) + 1e-12)

            ##################################################特征选择############################################
            # 分类器的rdstate
            radio_grid_rdstate = random.randint(1, 9999)  # 各个分类器的random state auc 63 acc 70 23
            dsfr_grid_rdstate = random.randint(1, 9999)
            radio_grid_rdstate= 5224
            dsfr_grid_rdstate= 8046
            #grid_rdstate = 1390  # 各个分类器的random state vegfr2 1204 sstr2 1715 mgmt 1204
            print('radio_grid_rdstate', radio_grid_rdstate)
            print('dsfr_grid_rdstate', dsfr_grid_rdstate)
            '''
            1:dsfr
            2:radiomics
            3:'dsfr+radiomics'
            4:dsfr  radiomics分数相加
            5.dsfr  dsfr+radiomics分数相加
            6.radiomics  dsfr+radiomics分数相加
            7 dsfr radiomics  dsfr+radiomics 分数相加

            8 radiomics *+lr
            '''
            key_word_model = ['dsfr', 'radiomics', 'dsfr+radiomics']
            show_choose = 4
            alpha_radio = 0.0004933
            alpha_dsfr = 0.001127
            pers_chi = 40
            xuhao_radio = np.array([[it for it in range(radiofeature_out.shape[1])]])
            xuhao_dsfr = np.array([[it for it in range(dsfrfeature_out.shape[1])]])

            inner_radio_feature_choose, out_radio_feature_choose,xuhao_radio = choose_feature(radiofeature,label=label,feature_out=radiofeature_out,xuhao = xuhao_radio,uselasso=True,alpha=alpha_radio,random_state=radio_grid_rdstate)#0.000795
            inner_radio_feature_choose, out_radio_feature_choose,xuhao_radio = choose_feature(inner_radio_feature_choose,label=label,feature_out=out_radio_feature_choose,xuhao = xuhao_radio,useSelectPercentile=True, SelectPercentile_fc=3, pers_chi=60,random_state=radio_grid_rdstate)#95


            inner_dsfr_feature_choose, out_dsfr_feature_choose,xuhao_dsfr = choose_feature(dsfrfeature,label=label,feature_out=dsfrfeature_out,uselasso=True,alpha=alpha_dsfr,xuhao = xuhao_dsfr,random_state=dsfr_grid_rdstate)#0.000955  20
            inner_dsfr_feature_choose, out_dsfr_feature_choose,xuhao_dsfr = choose_feature(inner_dsfr_feature_choose,label=label,feature_out=out_dsfr_feature_choose,useSelectPercentile=True,xuhao = xuhao_dsfr,SelectPercentile_fc=5, pers_chi=60,random_state=dsfr_grid_rdstate)#3  42
            # fc_random=  66613  #SelectPercentile_fc=5的随机数，需要跳进去原始代码里面设置

            for it,icc_feature in enumerate(icc_excel.Feature):
                inner_radio_feature_choose=np.delete(inner_radio_feature_choose,np.where(xuhao_radio[0,:]==icc_feature),axis=1)
                out_radio_feature_choose=np.delete(out_radio_feature_choose, np.where(xuhao_radio[0, :] == icc_feature), axis=1)
                xuhao_radio=np.delete(xuhao_radio, np.where(xuhao_radio[0, :] == icc_feature), axis=1)
                if it >273:
                    break
            '''                                                                                                 
           choose_feature(feature,feature_out=None,select_first=True,usepca = False,usekernelpca = False,uselda = False,useUmap = False,
                   usestatics = False,uselasso = False,alpha = 0.0008,uselassocv = False,useSelectPercentile=False,SelectPercentile_fc=2,pers_chi=5):
                               '''
            print('radio choose feature:{} \ndsfr choose feature:{} '.format(inner_radio_feature_choose.shape[1],inner_dsfr_feature_choose.shape[1]))

            if 'dsfr+radiomics' in key_word_model:
                xuhao_d_r = np.array([[it for it in range(dsfrfeature_out.shape[1])]])
                inner_dsfr_feature_choose1, out_dsfr_feature_choose1,_ = choose_feature(dsfrfeature,label=label,xuhao= xuhao_d_r,feature_out=dsfrfeature_out,uselasso=True,alpha=0.0191,random_state=dsfr_grid_rdstate)  # 0.0191
                inner_r_d_f = np.hstack((radiofeature, inner_dsfr_feature_choose1))
                if test_outside:
                    out_r_d_f = np.hstack((radiofeature_out, out_dsfr_feature_choose1))
                inner_r_d_f_choose, out_r_d_f_choose,_ = [inner_r_d_f, out_r_d_f, 0.0056]
            # 分折
            num_subjects = inner_radio_feature_choose.shape[0]
            cvrdstate = 8116
            print('cvrdstate', cvrdstate)
            if holdout:
                sfolder = KFold(n_splits=num_subjects, random_state=cvrdstate, shuffle=True)
            else:
                sfolder = StratifiedKFold(n_splits=k_fold, random_state=cvrdstate, shuffle=True)
                # ps:shuffle=True时，random_state才有用
                # sfolder = StratifiedKFold(n_splits=k_fold, shuffle=False)



            # 开搞
            # 构建储存结果的容器
            label_CV = []
            id_CV = []
            dsfrprob_lr = []
            radioprob_lr = []
            d_r_prob_lr = []
            radio_important = []
            dsfr_important = []

            if test_outside:
                dsfrlabel_CV_out = []
                dsfrprob_lr_out = []

                radiolabel_CV_out = []
                radioprob_lr_out = []

                d_r_prob_lr_out = []

            # 固定参数 ============================================================================================
            final_lr_penalty = 'l2'
            final_lr_C = 0.1
            final_lr_max_iter = 400
            final_lr_tol = 0.001

            fold_flag = 1  # 记录交叉验证跑到第几折了
            for train, test in sfolder.split(inner_dsfr_feature_choose, label):
                print('doing with fold ', fold_flag)
                #print(test)

                fold_flag = fold_flag + 1
                dsfrfeature_train, label_train = [inner_dsfr_feature_choose[train], label[train]]
                radiofeature_train, label_train = [inner_radio_feature_choose[train], label[train]]

                dsfrfeature_test, label_test = [inner_dsfr_feature_choose[test], label[test]]
                radiofeature_test, label_test = [inner_radio_feature_choose[test], label[test]]

                id_CV = id_CV + list(id_all[test])
                current_fold_train_id = list(id_all[train])
                current_fold_test_id = list(id_all[test])

                if 'dsfr+radiomics' in key_word_model:
                    inner_r_d_f_train, _ = [inner_r_d_f_choose[train], label[train]]
                    inner_r_d_f_test, _ = [inner_r_d_f_choose[test], label[test]]

                if test_outside:
                    dsfrfeature_out_test = copy.deepcopy(out_dsfr_feature_choose)
                    radiofeature_out_test = copy.deepcopy(out_radio_feature_choose)
                    if 'dsfr+radiomics' in key_word_model:
                        out_r_d_f_test = copy.deepcopy(out_r_d_f_choose)

                if use_subject_decrease:
                    if decrease_method == 3:
                        current_fold_train_id_index = []
                        dsfrfeature_train, label_train1, current_fold_train_id1 = sub_undersampled_use_clustering(dsfrfeature_train, label_train, current_fold_train_id)

                        current_fold_train_id_index = [int(it) for it,pth_id in enumerate(current_fold_train_id) if pth_id in current_fold_train_id1]

                        radiofeature_train, label_train, current_fold_train_id = [radiofeature_train[np.array(current_fold_train_id_index)],label_train1,current_fold_train_id1]
                        if 'dsfr+radiomics' in key_word_model:
                            inner_r_d_f_train, label_train, current_fold_train_id = [inner_r_d_f_train[np.array(current_fold_train_id_index)],label_train1,current_fold_train_id1]
                # 标准化 # 将上述得到的scale参数应用至测试数据
                StandardScaler = preprocessing.StandardScaler()  #
                dsfrfeature_train = StandardScaler.fit_transform(dsfrfeature_train)
                dsfrfeature_test = StandardScaler.transform(dsfrfeature_test)

                if test_outside:
                    dsfrfeature_out_test = StandardScaler.transform(dsfrfeature_out_test)

                StandardScaler = preprocessing.StandardScaler()
                radiofeature_train = StandardScaler.fit_transform(radiofeature_train)
                radiofeature_test = StandardScaler.transform(radiofeature_test)

                # 外部数据
                if test_outside:
                    radiofeature_out_test = StandardScaler.transform(radiofeature_out_test)

                # 特征筛选之后,附加辅助变量 ========================================================================
                # 这里比较特殊,因为是在交叉验证里面,所以需要先将特征分开为训练集和测试集,之后再附加
                # 这里的代码之后再写(大雾)@.@

                # 分类器设置 ======================================================================================
                # Logistics ---------------------------------------------------------------------------------------
                if UseLR:
                    if random_grid:
                        if 'dsfr' in key_word_model:
                            clf = LogisticRegression(penalty=final_lr_penalty, C=final_lr_C, max_iter=final_lr_max_iter,
                                                     tol=final_lr_tol, solver='lbfgs', warm_start=True,
                                                     random_state=dsfr_grid_rdstate)

                            clf.fit(dsfrfeature_train, label_train)

                            # 预测测试集
                            y_prob = clf.predict_proba(dsfrfeature_test)
                            dsfrprob_lr = dsfrprob_lr + list(y_prob[:, 1])
                            important = clf.coef_[0]
                            dsfr_important.append(list(important))
                            # 测试外验证
                            if test_outside:
                                y_prob = clf.predict_proba(dsfrfeature_out_test)
                                dsfrprob_lr_out.append(list(y_prob[:, 1]))

                        # key_word_model = ['dsfr', 'radiomics', 'dsfr+radiomics']  #
                        if 'radiomics' in key_word_model:
                            clf = LogisticRegression(penalty=final_lr_penalty, C=final_lr_C, max_iter=final_lr_max_iter,
                                                     tol=final_lr_tol, solver='lbfgs', warm_start=True,
                                                     random_state=radio_grid_rdstate)

                            clf.fit(radiofeature_train, label_train)
                            # 预测测试集
                            y_prob = clf.predict_proba(radiofeature_test)
                            radioprob_lr = radioprob_lr + list(y_prob[:, 1])
                            important = clf.coef_[0]
                            radio_important.append(list(important))
                            if test_outside:
                                y_prob = clf.predict_proba(radiofeature_out_test)
                                radioprob_lr_out.append(list(y_prob[:, 1]))

                        if 'dsfr+radiomics' in key_word_model:
                            clf = LogisticRegression(penalty=final_lr_penalty, C=final_lr_C, max_iter=final_lr_max_iter,
                                                     tol=final_lr_tol, solver='lbfgs', warm_start=True,
                                                     random_state=radio_grid_rdstate)
                            clf.fit(inner_r_d_f_train, label_train)
                            # 预测测试集
                            y_prob = clf.predict_proba(inner_r_d_f_test)
                            d_r_prob_lr = d_r_prob_lr + list(y_prob[:, 1])
                            if test_outside:
                                if 'dsfr+radiomics' in key_word_model:
                                    y_prob = clf.predict_proba(out_r_d_f_test)
                                    d_r_prob_lr_out.append(list(y_prob[:, 1]))

                # 保存label --------------------------------------------------------------------------------
                label_CV = label_CV + list(label_test)

            if show_choose == 1:
                prob_lr = dsfrprob_lr
                if test_outside:
                    prob_lr_out = dsfrprob_lr_out
            elif show_choose == 2:
                prob_lr = radioprob_lr
                if test_outside:
                    prob_lr_out = radioprob_lr_out
            elif show_choose == 3:
                prob_lr = d_r_prob_lr
                if test_outside:
                    prob_lr_out = d_r_prob_lr_out
            elif show_choose == 4:
                prob_lr = (np.array(dsfrprob_lr) + np.array(radioprob_lr)) / 2
                if test_outside:
                    prob_lr_out = (np.array(dsfrprob_lr_out) + np.array(radioprob_lr_out)) / 2
            elif show_choose == 5:
                prob_lr = (np.array(dsfrprob_lr) + np.array(d_r_prob_lr)) / 2
                if test_outside:
                    prob_lr_out = (np.array(dsfrprob_lr_out) + np.array(d_r_prob_lr_out)) / 2
            elif show_choose == 6:
                prob_lr = (np.array(radioprob_lr) + np.array(d_r_prob_lr)) / 2
                if test_outside:
                    prob_lr_out = (np.array(radioprob_lr_out) + np.array(d_r_prob_lr_out)) / 2
            elif show_choose == 7:
                prob_lr = (np.array(radioprob_lr) + np.array(d_r_prob_lr) + np.array(dsfrprob_lr)) / 3
                if test_outside:
                    prob_lr_out = (np.array(radioprob_lr_out) + np.array(d_r_prob_lr_out) + np.array(
                        dsfrprob_lr_out)) / 3
            elif show_choose == 8:
                prob_lr = (np.array(radioprob_lr) + np.array(radioprob_lr)) / 2
                if test_outside:
                    prob_lr_out = (np.array(radioprob_lr_out) + np.array(radioprob_lr_out)) / 2

            # 计算&保存内部数据指标
            if UseLR:
                acc_lr, sen_lr, spc_lr, AUC_lr, bst_lr = getAccSenSpcAuc(label_CV, prob_lr)
                # exc_name = r'E:\workspace\BC\experience\result\cut_result\2_01\HER2_2_01_important_radio.xlsx'
                # wb = workbook.Workbook()
                # ws = wb.active
                # ws.append(['Feature', 'Score'])
                # radio_important = list(np.mean(np.array(radio_important), axis=0))
                # sorted_id = sorted(range(len(radio_important)), key=lambda k: abs(radio_important[k]), reverse=True)
                # real_sorted_id = xuhao_radio[0][sorted_id]
                # print('元素索引序列：', sorted_id)
                # # summarize feature importance
                # for i, v in enumerate(sorted_id):
                #     print('Feature: %0d, Score: %.5f' % (xuhao_radio[0][v], radio_important[v]))
                #     ws.append([xuhao_radio[0][v], radio_important[v]])
                # wb.save(exc_name)

                # exc_name = r'E:\workspace\BC\experience\result\cut_result\2_01\HER2_2_01_inner_radio.xlsx'
                # wb = workbook.Workbook()  # 创建Excel对象
                # ws = wb.active  # 获取当前正在操作的表对象
                # ws.append(['id', 'true_label', 'score_N2'])
                # for it in range(len(id_CV)):
                #     ws.append([id_CV[it], label_CV[it], prob_lr[it]])
                # wb.save(exc_name)

            # 计算外验证指标，注意，计算指标时候使用内部数据的最佳阈值!!!
            if test_outside:
                label_CV_out = list(label_out)
                if UseLR:
                    prob_lr_out = list(np.mean(np.array(prob_lr_out), axis=0))
                    # acc_lr_out, sen_lr_out, spc_lr_out, AUC_lr_out, bst_lr_out = getAccSenSpcAuc(label_CV_out, prob_lr_out, pre_bestthresold=bst_lr)
                    acc_lr_out, sen_lr_out, spc_lr_out, AUC_lr_out, bst_lr_out = getAccSenSpcAuc(label_CV_out,prob_lr_out)
                    # exc_name = r'E:\workspace\BC\experience\result\cut_result\2_01\HER2_2_01_out_radio.xlsx.xlsx'
                    # wb = workbook.Workbook()  # 创建Excel对象
                    # ws = wb.active  # 获取当前正在操作的表对象
                    # ws.append(['id', 'true_label', 'score_N2'])
                    # for it in range(len(id_all_out)):
                    #     ws.append([id_all_out[it], label_CV_out[it], prob_lr_out[it]])
                    # wb.save(exc_name)

            filename = savepath + sep + 'result.txt'
            f = open(filename, 'a+')
            writewords = str(test1) + r'===========================================================' + '\n' + \
                         r' $  holdout: ' + str(
                holdout) + ', distill_mode : ' + distill_mode + ',' + \
                         '  cvrdstate: ' + str(cvrdstate) +'\n'+ \
                         'radio_grid_rdstate: ' + str(radio_grid_rdstate) + '\n' + \
                         '  dsfr_grid_rdstate: ' + str(dsfr_grid_rdstate) + '\n' + \
                        'alpha_radio' + str(alpha_radio) + ',' + \
                        '   alpha_dsfr' + str(alpha_dsfr) + '\n' + \
                        'pers_chi' + str(pers_chi) + '\n' + \
                         '  k_fold:' + str(k_fold) + '\n'
            if UseLR:
                writewords = writewords + r' @ lr   @ AUC: ' + str('%03f' % AUC_lr) + \
                             ',Acc: ' + str('%03f' % acc_lr) + \
                             ',Sen: ' + str('%03f' % sen_lr) + \
                             ',Spc: ' + str('%03f' % spc_lr) + \
                             ',Best_thresold: ' + str('%03f' % bst_lr) + '\n'


            print(writewords)
            f.write(writewords)
            f.close()

            # 保存外验证指标
            if test_outside:
                filename = savepath + sep + 'result.txt'
                f = open(filename, 'a+')
                writewords = '\n' + r'* outside *' + '\n'
                if UseLR:
                    writewords = writewords + r' @ lr   @ outAUC: ' + str('%03f' % AUC_lr_out) + \
                                 ',Acc: ' + str('%03f' % acc_lr_out) + \
                                 ',Sen: ' + str('%03f' % sen_lr_out) + \
                                 ',Spc: ' + str('%03f' % spc_lr_out) + \
                                 ',Best_thresold: ' + str('%03f' % bst_lr_out) + '\n'

                print(writewords)
                f.write(writewords)
                f.close()

            # 保存预测结果(用来做ROC的data)、label、id
            if save_prob:
                filename = savepath + sep + 'rocdata.txt'
                f = open(filename, 'a+')
                writewords = 'test' + str(test1) + r'=============================================:' + '\n' + \
                             r' @ID  ' + str(list(id_CV)) + '\n' + \
                             r' @Label  ' + str(label_CV) + '\n'

                if UseLR:
                    writewords = writewords + r' @lr  ' + str(prob_lr) + '\n'

                f.write(writewords)
                f.close()

                if test_outside:
                    filename = savepath + sep + 'rocdata_out.txt'
                    f = open(filename, 'a+')
                    writewords = 'test' + str(test1) + r'=============================================:' + '\n' + \
                                 r' @ID  ' + str(list(id_all_out)) + '\n' + \
                                 r' @Label  ' + str(label_CV_out) + '\n'
                    if UseLR:
                        writewords = writewords + r' @lr  ' + str(prob_lr_out) + '\n'

                    f.write(writewords)
                    f.close()

            # 保存参数
            if random_grid:
                filename = savepath + sep + 'param.txt'
                f = open(filename, 'a+')
                f.write('test：' + str(test1) + '=====================================================' + '\n')

                if UseLR:
                    f.write('final_lr_penalty  =' + str(final_lr_penalty) + '\n')
                    f.write('final_lr_C  =' + str(final_lr_C) + '\n')
                    f.write('final_lr_max_iter  =' + str(final_lr_max_iter) + '\n')
                    f.write('final_lr_tol  =' + str(final_lr_tol) + '\n')
                f.close()






